import pandas as pd
import re
import traceback
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties

# 設定檔案名稱
input_file = 'Sunplus_Yield_control_table.xlsx'
output_file = 'yield_trend_6.xlsx'
sheet_name = 'QAL642E LFBGA 487B'
columns_to_keep = "B, C, D, F, G, S, T"

try:
    # 1️⃣ 讀取 Excel，篩選特定欄位（用欄位位置），跳過第一列
    df = pd.read_excel(input_file, sheet_name=sheet_name, usecols=columns_to_keep, skiprows=1)
    df.to_excel('yield_trend_a.xlsx')

    # 2️⃣ 新增 RT rate 欄位
    df["RT rate"] = None
    df.to_excel('yield_trend_b.xlsx')

    # ...已移除空值與型態檢查...

    # 3️⃣ 修改 Station 名稱
    def modify_ft(station, pgm_name):
        if station == "FT":
            match = re.search(r"f(\d+)", pgm_name)
            if match:
                return f"FT{match.group(1)}"
        return station

    df["Station"] = df.apply(lambda row: modify_ft(row["Station"], row["PGM Name"]), axis=1)
    df.to_excel('yield_trend_c.xlsx')

    # 4️⃣ 計算 RT rate（修正：避免 None 與 int 比較）
    rt_rate = None
    rt_start_idx = None

    for idx in df.index:
        station = str(df.at[idx, "Station"])

        if station.startswith("FT"):
            rt_rate = 0
            rt_start_idx = idx

        elif re.match(r"R(\d+)", station):
            match = re.match(r"R(\d+)", station)
            if match:
                r_value = int(match.group(1))
                if rt_rate is None:
                    rt_rate = r_value
                else:
                    rt_rate = max(rt_rate, r_value)

        elif station == "Total" and rt_start_idx is not None:
            df.loc[rt_start_idx:idx, "RT rate"] = rt_rate
            rt_rate = None
            rt_start_idx = None

    df.to_excel('yield_trend_d.xlsx')

    # 5️⃣ 刪除包含 NaN 的列
    df_cleaned = df.dropna()
    df_cleaned.to_excel('yield_trend_e.xlsx')

    # 6️⃣ 分類 FT1, FT2, FT3 到不同 Sheet，並收集統計資料
    stats = []
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for ft_group in df_cleaned["Station"].unique():
            if ft_group.startswith("FT"):
                ft_df = df_cleaned[df_cleaned["Station"] == ft_group]
                ft_df.to_excel(writer, sheet_name=ft_group, index=False)
                # 統計分析
                avg = ft_df["Overall Yield"].mean()
                std = ft_df["Overall Yield"].std()
                minv = ft_df["Overall Yield"].min()
                maxv = ft_df["Overall Yield"].max()
                stats.append({
                    "Station": ft_group,
                    "平均": avg,
                    "標準差": std,
                    "最大值": maxv,
                    "最小值": minv
                })
        # 匯出統計摘要到 Summary Sheet
        import numpy as np
        summary_df = pd.DataFrame(stats)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # 7️⃣ 調整 Excel 欄寬、8️⃣ 統一 RT rate Y 軸高度、9️⃣ 加入圖表
    wb = load_workbook(output_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    max_rt_rate = df_cleaned["RT rate"].dropna().max()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_headers = [str(cell.value) for cell in ws[1]]
        # 只處理包含 Lot# 欄位的分頁（略過 Summary Sheet）
        if "Lot#" not in raw_headers:
            print(f"[略過分頁] {sheet_name}，因為缺少 Lot# 欄位")
            continue

        def find_col_exact(name):
            if name in raw_headers:
                return raw_headers.index(name) + 1
            else:
                print(f"❌ 找不到欄位: {name}")
                print("[欄位名稱清單]", raw_headers)
                print(f"請修改程式中的欄位名稱設定，或直接用下列名稱之一：{raw_headers}")
                raise ValueError(f"請確認欄位名稱設定！")

        lot_col = find_col_exact("Lot#")
        first_pass_col = find_col_exact("First Pass Yield")
        overall_col = find_col_exact("Overall Yield")
        rt_rate_col = find_col_exact("RT rate")
        last_row = ws.max_row

        # 折線圖
        combo_chart = LineChart()
        combo_chart.title = ""
        combo_chart.x_axis.title = "Lot#"
        combo_chart.y_axis.title = "Yield (%)"

        x_values = Reference(ws, min_col=lot_col, min_row=2, max_row=last_row)

        for col_index in [first_pass_col, overall_col]:
            y_values = Reference(ws, min_col=col_index, min_row=1, max_row=last_row)
            combo_chart.add_data(y_values, titles_from_data=True)

        # 讓折線圖恢復稜角（不平滑）
        for s in combo_chart.series:
            s.smooth = False

        combo_chart.set_categories(x_values)
        # 讓每個 Lot# 都顯示在 X 軸
        combo_chart.x_axis.tickLblSkip = 1

        # 加標準線
        std_line_values = [0.98] * (last_row - 1)
        for i, val in enumerate(std_line_values, start=2):
            ws.cell(row=i, column=overall_col + 2, value=val)

        std_line = Reference(ws, min_col=overall_col + 2, min_row=2, max_row=last_row)
        std_series = Series(std_line, title="標準線 (0.98)")
        std_series.graphicalProperties.line.solidFill = "808080"
        std_series.graphicalProperties.line.dashStyle = "sysDash"
        combo_chart.append(std_series)

        # 柱狀圖 RT rate
        bar_chart = BarChart()
        bar_chart.y_axis.title = "RT rate"
        bar_chart.y_axis.axId = 200
        bar_chart.y_axis.majorGridlines = None

        y_values = Reference(ws, min_col=rt_rate_col, min_row=1, max_row=last_row)
        bar_chart.add_data(y_values, titles_from_data=True)
        bar_chart.set_categories(x_values)

        combo_chart.y_axis.crosses = "max"
        combo_chart += bar_chart

        bar_chart.y_axis.scaling.min = 0
        bar_chart.y_axis.scaling.max = max_rt_rate * 2.0

        # 淡化格線
        gray_gridlines = ChartLines()
        gray_gridlines.spPr = GraphicalProperties()
        gray_gridlines.spPr.ln = LineProperties(solidFill=ColorChoice(prstClr="ltGray"))
        combo_chart.y_axis.majorGridlines = gray_gridlines

        # 放大圖表
        combo_chart.width = 24
        combo_chart.height = 12

        combo_chart.legend.position = "t"
        combo_chart.legend.layout = None
        combo_chart.legend.overlay = False

        # 插入圖表
        ws.add_chart(combo_chart, "K5")

    wb.save(output_file)
    print(f"✅ {output_file} 已成功儲存，圖例已移至圖表上方外部水平排列！")

except FileNotFoundError:
    print("❌ 找不到原始檔案，請檢查檔案名稱和路徑。")
except ValueError as e:
    print(f"❌ 發生錯誤: {e}")
except Exception as e:
    print(f"❌ 發生未知錯誤: {e}")
    traceback.print_exc()
