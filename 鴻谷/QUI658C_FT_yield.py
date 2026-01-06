import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties

# 設定檔案名稱
input_file = 'Sunplus_Yield_control_table.xlsx'
output_file = 'QUI658C_FT_yield_trend.xlsx'
sheet_name = 'QUI658C 128MCM(LQFP)'

# 指定要保留的欄位
columns_to_keep = "B, C, D, F, G, S, T"

try:
    # **1️⃣ 讀取 Excel，篩選特定欄位，跳過第一列**
    df = pd.read_excel(input_file, sheet_name=sheet_name, usecols=columns_to_keep, skiprows=1)
    df.to_excel('yield_trend_a.xlsx') 

    # **2️⃣ 新增 RT rate 欄位**
    df["RT rate"] = None  # 預設值
    df.to_excel('yield_trend_b.xlsx') 


    # **3️⃣ 解析 PGM Name，修改 FT 為 FT1、FT2...**
    def modify_ft(station, pgm_name):
        """ 如果 Station 是 FT，則從 PGM Name 中提取 f 後的數字，變成 FT1、FT2... """
        if station == "FT":
            match = re.search(r"f(\d+)", pgm_name)
            if match:
                return f"FT{match.group(1)}"
        return station

    df["Station"] = df.apply(lambda row: modify_ft(row["Station"], row["PGM Name"]), axis=1)
    df.to_excel('yield_trend_c.xlsx')

    # **4️⃣ 計算 RT rate**
    rt_rate = None

    for idx in df.index:
        station = str(df.at[idx, "Station"])

        if station.startswith("FT"):
            rt_rate = 0
            rt_start_idx = idx

        elif re.match(r"R(\d+)", station):
            rt_rate = max(rt_rate, int(re.match(r"R(\d+)", station).group(1)))

        elif station == "Total":
            df.loc[rt_start_idx:idx, "RT rate"] = rt_rate
            rt_rate = None

    df.to_excel('yield_trend_d.xlsx')        

    # **5️⃣ 刪除包含 NaN 的列**   
    df_cleaned = df.dropna()
    df_cleaned.to_excel('yield_trend_e.xlsx')

    # **6️⃣ 分類 FT1, FT2, FT3 到不同的 Sheet**
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for ft_group in df_cleaned["Station"].unique():
            if ft_group.startswith("FT"):
                df_cleaned[df_cleaned["Station"] == ft_group].to_excel(writer, sheet_name=ft_group, index=False)

    # **7️⃣ 調整 Excel 欄位寬度**
    wb = load_workbook(output_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    
    # **8️⃣ 統一 RT rate Y 軸高度** 
    max_rt_rate = max(df_cleaned["RT rate"])

    # **9️⃣ 為每個 FT Sheet 加入趨勢圖**
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        lot_col = headers.index("Lot#") + 1
        first_pass_col = headers.index("First Pass Yield") + 1
        overall_col = headers.index("Overall Yield") + 1
        rt_rate_col = headers.index("RT rate") + 1
        last_row = ws.max_row

         # **建立折線圖**
        combo_chart = LineChart()
        combo_chart.title = ""
        combo_chart.x_axis.title = "Lot#"
        combo_chart.y_axis.title = "Yield (%)"

        x_values = Reference(ws, min_col=lot_col, min_row=2, max_row=last_row)

        for col_index in [first_pass_col, overall_col]:
            y_values = Reference(ws, min_col=col_index, min_row=1, max_row=last_row)
            combo_chart.add_data(y_values, titles_from_data=True)

        combo_chart.set_categories(x_values)
        
        # **建立標準線 (Target Line)**
        std_line_values = [0.98] * (last_row - 1)
        for i, val in enumerate(std_line_values, start=2):
            ws.cell(row=i, column=overall_col + 2, value=val)

        std_line = Reference(ws, min_col=overall_col + 2, min_row=2, max_row=last_row)
        std_series = Series(std_line, title="標準線 (0.98)")
        std_series.graphicalProperties.line.solidFill = "808080"  # 灰色
        std_series.graphicalProperties.line.dashStyle = "sysDash"  # 虛線
        combo_chart.append(std_series)  # ✅ 與折線圖對齊！

         # **建立柱狀圖**
        bar_chart = BarChart()
        bar_chart.y_axis.title = "RT rate"
        bar_chart.y_axis.axId = 200
        bar_chart.y_axis.majorGridlines = None

        y_values = Reference(ws, min_col=rt_rate_col, min_row=1, max_row=last_row)
        bar_chart.add_data(y_values, titles_from_data=True)
        bar_chart.set_categories(x_values)

        combo_chart.y_axis.crosses = "max"
        combo_chart += bar_chart

        bar_chart.y_axis.scaling.min = 0
        bar_chart.y_axis.scaling.max = max_rt_rate * 2.0

        # **淡化主要格線**
        gray_gridlines = ChartLines()
        gray_gridlines.spPr = GraphicalProperties()
        gray_gridlines.spPr.ln = LineProperties(solidFill=ColorChoice(prstClr="ltGray"))
        combo_chart.y_axis.majorGridlines = gray_gridlines

        # **放大圖表**
        combo_chart.width = 24
        combo_chart.height = 12

        # 將圖例放到圖表上方，橫向排列，並設在圖表外部
        combo_chart.legend.position = "t"
        combo_chart.legend.layout = None
        combo_chart.legend.overlay = False

        # **📌 將圖表插入位置往左移一欄**
        ws.add_chart(combo_chart, "K5")

    wb.save(output_file)
    print(f"✅ {output_file} 已成功儲存，圖例已移至圖表上方外部水平排列！")

except FileNotFoundError:
    print("❌ 找不到原始檔案，請檢查檔案名稱和路徑。")
except ValueError as e:
    print(f"❌ 發生錯誤: {e}")
except Exception as e:
    print(f"❌ 發生未知錯誤: {e}")
