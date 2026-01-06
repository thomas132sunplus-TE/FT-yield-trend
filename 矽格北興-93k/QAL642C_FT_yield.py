import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.shapes import GraphicalProperties

# 設定檔案名稱
input_file = 'Sunplus_Yield_control_table.xlsx'
output_file = 'QAL642C_FT_yield_trend.xlsx'
sheet_name = 'QAL642C LFBGA 487B'

# 指定要保留的欄位
columns_to_keep = "B, C, D, F, G, S, T"

try:
    df = pd.read_excel(input_file, sheet_name=sheet_name, usecols=columns_to_keep, skiprows=1)
    df["RT rate"] = None

    def modify_ft(station, pgm_name):
        if station == "FT":
            match = re.search(r"f(\d+)", pgm_name)
            if match:
                return f"FT{match.group(1)}"
        return station

    df["Station"] = df.apply(lambda row: modify_ft(row["Station"], row["PGM Name"]), axis=1)

    rt_rate = None
    for idx in df.index:
        station = str(df.at[idx, "Station"])
        if station.startswith("FT"):
            rt_rate = 0
            rt_start_idx = idx
        elif re.match(r"R(\d+)", station):
            rt_rate = max(rt_rate, int(re.match(r"R(\d+)", station).group(1)))
        elif station == "Total":
            df.loc[rt_start_idx:idx, "RT rate"] = rt_rate
            rt_rate = None

    df_cleaned = df.dropna()

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for ft_group in df_cleaned["Station"].unique():
            if ft_group.startswith("FT"):
                df_cleaned[df_cleaned["Station"] == ft_group].to_excel(writer, sheet_name=ft_group, index=False)

    wb = load_workbook(output_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    max_rt_rate = max(df_cleaned["RT rate"])

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        lot_col = headers.index("Lot#") + 1
        first_pass_col = headers.index("First Pass Yield") + 1
        overall_col = headers.index("Overall Yield") + 1
        rt_rate_col = headers.index("RT rate") + 1
        last_row = ws.max_row

        combo_chart = LineChart()
        combo_chart.title = ""
        combo_chart.x_axis.title = "Lot#"
        combo_chart.y_axis.title = "Yield (%)"

        x_values = Reference(ws, min_col=lot_col, min_row=2, max_row=last_row)

        for col_index in [first_pass_col, overall_col]:
            y_values = Reference(ws, min_col=col_index, min_row=1, max_row=last_row)
            combo_chart.add_data(y_values, titles_from_data=True)

        combo_chart.set_categories(x_values)

        std_line_values = [0.98] * (last_row - 1)
        for i, val in enumerate(std_line_values, start=2):
            ws.cell(row=i, column=overall_col + 2, value=val)

        std_line = Reference(ws, min_col=overall_col + 2, min_row=2, max_row=last_row)
        std_series = Series(std_line, title="標準線 (0.98)")
        std_series.graphicalProperties.line.solidFill = "808080"
        std_series.graphicalProperties.line.dashStyle = "sysDash"
        combo_chart.append(std_series)

        bar_chart = BarChart()
        bar_chart.y_axis.title = "RT rate"
        bar_chart.y_axis.axId = 200
        bar_chart.y_axis.majorGridlines = None

        y_values = Reference(ws, min_col=rt_rate_col, min_row=1, max_row=last_row)
        bar_chart.add_data(y_values, titles_from_data=True)
        bar_chart.set_categories(x_values)

        combo_chart.y_axis.crosses = "max"
        combo_chart += bar_chart

        bar_chart.y_axis.scaling.min = 0
        bar_chart.y_axis.scaling.max = max_rt_rate * 2.0

        gray_gridlines = ChartLines()
        gray_gridlines.spPr = GraphicalProperties()
        gray_gridlines.spPr.ln = LineProperties(solidFill=ColorChoice(prstClr="ltGray"))
        combo_chart.y_axis.majorGridlines = gray_gridlines

        combo_chart.width = 24
        combo_chart.height = 12

        # 將圖例放到圖表上方，橫向排列，並設在圖表外部
        combo_chart.legend.position = "t"
        combo_chart.legend.layout = None
        combo_chart.legend.overlay = False

        ws.add_chart(combo_chart, "K5")

    wb.save(output_file)
    print(f"✅ {output_file} 已成功儲存，圖例已移至圖表上方外部水平排列！")

except FileNotFoundError:
    print("❌ 找不到原始檔案，請檢查檔案名稱和路徑。")
except ValueError as e:
    print(f"❌ 發生錯誤: {e}")
except Exception as e:
    print(f"❌ 發生未知錯誤: {e}")
